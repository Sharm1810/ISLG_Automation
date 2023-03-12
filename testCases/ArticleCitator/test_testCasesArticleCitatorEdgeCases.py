import sys
from operator import contains
from telnetlib import EC

import clipboard
import pytest
import time
import json

import softest
from IPython.utils import data
from py._builtin import text
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.ui import Select

from pageObjects.ArticleCitator import ArticleCitator
from pageObjects.LoginPage import LoginPage
from pageObjects.SubjectNavigator import SubjectNavigator
from utilities.readProperties import ReadConfig
from utilities.customLogger import LogGen
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from bs4 import BeautifulSoup
import requests
import string
import pyperclip
import pandas as pd
import random
import csv
import openpyxl


@pytest.mark.usefixtures("setup")
class Test_test_testCasesArticleCitatorEdgeCases(softest.TestCase):
    logger = LogGen.loggen()

    ####Paul's requirement####

    @pytest.mark.skip(reason="None")
    def test_expandAll(self):
        self.logger.info("****TestCase AC-018 - Expand All***")
        self.logger.info("*****Login Successful****")
        self.logger.info("**** Article Citator testing *****")
        self.navigator = ArticleCitator(self.driver)
        self.navigator.clickOnArticleCitatormenu()
        time.sleep(2)
        elems = self.driver.find_elements_by_xpath("//div[@class='ACCardData']//div//div//a")
        for elem in elems:
            print(elem.get_attribute("data-searchtitle"))
            self.driver.execute_script("arguments[0].click();", elem)
            time.sleep(2)
            viewAllProvisions = self.driver.find_element(By.XPATH, "//div[@class='document__bar']//div//div//p//a")
            self.driver.execute_script("arguments[0].click();", viewAllProvisions)
            time.sleep(2)
            # pages = self.driver.find_elements(By.XPATH, "//li[@class='dynamicpagecount active']//a")
            # for page in pages:
            provisionExtract = self.driver.find_elements(By.PARTIAL_LINK_TEXT, "provision extract")
            if provisionExtract:
                for provision in provisionExtract:
                    self.driver.execute_script("arguments[0].click();", provision)
                    time.sleep(3)
                    try:
                        provisionText = self.driver.find_element(By.XPATH,
                                                                 "//div[@id='popup-prv-extract']//div//div[2]//div//div[2]")
                        if len(provisionText.text) > 0:
                            print("Text displayed")
                        else:
                            print("Elem found but length is not correct, handle this case.")
                    except NoSuchElementException:
                        print("Text is not displayed")
                    closebutton = self.driver.find_element(By.XPATH,
                                                           "//*[@id='popup-prv-extract']/div[1]/div[1]/button")
                    self.driver.execute_script("arguments[0].click();", closebutton)
                    time.sleep(2)
                pages = self.driver.find_element(By.XPATH,
                                                 "//div[@class='divProvisionRender open']//nav//ul//li[@class='dynamicpagecount active']//a")
                if pages:
                    self.driver.execute_script("arguments[0].click();", pages)
                    time.sleep(2)
                    # a = len(pages)
                    # for i in range(1, a):
                    #     print(i, end=" ")

                    # pagination = len(self.driver.find_elements(By.TAG_NAME, 'li'))
                    # print(pagination)

                    # pagination = self.driver.find_element(By.XPATH, "//li[@class='next']//a")

                # pageNav = self.driver.find_element(By.XPATH, "//i[@class='fas fa-caret-right']")
                # topicText = pageNav.get_attribute('aria-hidden')
                # time.sleep(2)
                # if topicText == "true":
                #         self.driver.execute_script("arguments[0].click();", pageNav)
                #         time.sleep(2)
                # for p in range(10):
                #     self.driver.find_element(By.XPATH, "//i[@class='fas fa-caret-right']").click()
                #     time.sleep(3)

    @pytest.mark.skip(reason="None")
    def test_paragraphExcerpts(self):
        self.logger.info("****TestCase AC-018 - Expand All***")
        self.logger.info("*****Login Successful****")
        self.logger.info("**** Article Citator testing *****")
        self.navigator = ArticleCitator(self.driver)
        self.navigator.clickOnArticleCitatormenu()
        time.sleep(2)
        elems = self.driver.find_elements_by_xpath("//div[@class='ACCardData']//div//div//a")
        for elem in elems:
            print(elem.get_attribute("data-searchtitle"))
            self.driver.execute_script("arguments[0].click();", elem)
            time.sleep(1)
            # Clicks on See All
            seeAll = self.driver.find_elements(By.LINK_TEXT, 'See All')
            for linkSee in seeAll:
                self.driver.execute_script("arguments[0].click();", linkSee)
                time.sleep(1)
                lnks = self.driver.find_elements(By.XPATH, "//a[starts-with(@tagpara,'pa')]")
                for lnk in lnks:
                    self.logger.info(lnk.get_attribute("tagpara"))
                    time.sleep(2)
                    self.driver.execute_script("arguments[0].click();", lnk)
                    time.sleep(2)
                #        # lnk.click()
                #        # # time.sleep(2)
                #

    #@pytest.mark.skip(reason="None")
    def test_expand100(self):
        global provisionText
        path = "C:/Users/sharm/Desktop/provisionExtract.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        r = 1
        c = 1
        self.logger.info("****TestCase AC-018 - Expand All***")
        self.logger.info("*****Login Successful****")
        self.logger.info("**** Article Citator testing *****")
        self.navigator = ArticleCitator(self.driver)
        self.navigator.clickOnArticleCitatormenu()
        time.sleep(2)
        # for x in range(len(self.driver.find_elements(By.XPATH, "(//div[@class='ACCardData']//div//div//a[1])"))):
        #     #if x % 100 == 0:
        #         clickme = self.driver.find_elements(By.XPATH, "(//div[@class='ACCardData']//div//div//a")
        #         self.driver.execute_script("arguments[0].click();", clickme)
        #         time.sleep(2)
        #         print("100 done")
        #         self.navigator.clickOnArticleCitatormenu()
        links = self.driver.find_elements(By.XPATH, "//div[@class='ACCardData']//div//div//a")

        time.sleep(2)
        batch_size = 100
        num_batches = len(links)
        for batch_index in range(num_batches):
             start_index = batch_index * batch_size
             end_index = min(start_index + batch_size, len(links))
             # Click on links in the current batch
             for link_index in range(start_index, end_index):
                 link = links[link_index]
                 self.driver.execute_script("arguments[0].click();", link)
                 time.sleep(2)
                 print(link.get_attribute("data-searchtitle"))
                 viewAllProvisions = self.driver.find_element(By.XPATH, "//div[@class='document__bar']//div//div//p//a")
                 self.driver.execute_script("arguments[0].click();", viewAllProvisions)
                 time.sleep(2)
                 # pages = self.driver.find_elements(By.XPATH, "//li[@class='dynamicpagecount active']//a")
                 # for page in pages:
                 provisionExtract = self.driver.find_elements(By.PARTIAL_LINK_TEXT, "provision extract")
                 if provisionExtract:
                     for provision in provisionExtract:
                         self.driver.execute_script("arguments[0].click();", provision)
                         time.sleep(3)
                         try:
                             provisionText = self.driver.find_element(By.XPATH,
                                                                      "//div[@id='popup-prv-extract']//div//div[2]//div//div[2]")

                             if len(provisionText.text) > 0:
                                 print("Text displayed :  ", provisionText.text)

                             else:
                                 print("Elem found but length is not correct, handle this case.")
                         except NoSuchElementException:
                             notext = self.driver.find_element(By.XPATH, "//div[@id='popup-prv-extract']//div//div//div//span//span")
                             ntext = notext.text
                             print("Text is not displayed", ntext)
                             sheet.cell(row=r, column=c).value = ntext
                             r = r + 1
                             workbook.save(path)

                         closebutton = self.driver.find_element(By.XPATH,
                                                                "//*[@id='popup-prv-extract']/div[1]/div[1]/button")
                         self.driver.execute_script("arguments[0].click();", closebutton)
                         time.sleep(2)
             print("100 Links")
             self.driver.refresh()
             #Find all the links again after the page refreshes
             links = self.driver.find_elements_by_xpath("//div[@class='ACCardData']//div//div//a")
        #     # Check if we have reached the end of the links
             if end_index >= len(links):
                 break
        #








        # # len(self.driver.find_elements(By.XPATH, "//div[@class='ACCardData']//div//div//a"))
        # elems = self.driver.find_elements_by_xpath("(//div[@class='ACCardData']//div//div//a[1])")
        # #for elem in elems:
        # for x in range(len(elems)):
        #     if x % 100 ==0:
        #         print(elems.get_attribute("data-searchtitle"))
        #         self.driver.execute_script("arguments[0].click();", elems)
        #         time.sleep(2)

        # count = len(self.driver.find_elements(By.XPATH, "(//div[@class='ACCardData']//div//div//a)"))
        # print(count)
        # for x in range(count):
        #     if x % 100 == 0:
        #         print(x)
        #         self.driver.find_elements(By.XPATH, "//div[@class='ACCardData']//div//div//a[1]")[x].click()
        #             # print(elem.get_attribute("data-searchtitle"))
        #             #     self.driver.execute_script("arguments[0].click();", elem)
        #             #     time.sleep(2)
        #
        #         # self.navigator.clickOnArticleCitatormenu()
        #         # time.sleep(2)

        # elems = self.driver.find_elements_by_xpath("(//div[@class='ACCardData']//div//div//a[1])[position() < 101]")
        # for elem in elems:
        #     print(elem.get_attribute("data-searchtitle"))
        #     self.driver.execute_script("arguments[0].click();", elem)
        #     time.sleep(2)
        #     provisionExtract = self.driver.find_elements(By.PARTIAL_LINK_TEXT, "provision extract")
        #     if provisionExtract:
        #         for provision in provisionExtract:
        #             self.driver.execute_script("arguments[0].click();", provision)
        #             time.sleep(3)
        #             try:
        #                 provisionText = self.driver.find_element(By.XPATH,
        #                                                          "//div[@id='popup-prv-extract']//div//div[2]//div//div[2]")
        #                 if len(provisionText.text) > 0:
        #                     print("Text is displayed")
        #
        #                 else:
        #                     print("take care")
        #             except NoSuchElementException:
        #                 print("Text is not displayed")
        #
        #             closebutton = self.driver.find_element(By.XPATH,
        #                                                    "//*[@id='popup-prv-extract']/div[1]/div[1]/button")
        #             self.driver.execute_script("arguments[0].click();", closebutton)
        #             time.sleep(2)
        #
        #
        #
        #
        #
        #
        #
    @pytest.mark.skip(reason="None")
    def test_writetoExcel(self):
        path = "C:/Users/sharm/Desktop/provisionExtract.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        for r in range(1, 1000):
            for c in range(1, 2):
                sheet.cell(row=r, column=c).value = "welcome"
        workbook.save(path)

